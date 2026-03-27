from flask import Blueprint, send_file, abort
from flask_login import login_required, current_user
from app.models import OF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from datetime import datetime

export_bp = Blueprint("export", __name__, url_prefix="/export")


def get_of_or_403(of_id):
    of = OF.query.filter_by(id=of_id, client_id=current_user.id).first()
    if not of:
        abort(404)
    return of


@export_bp.route("/of/<int:of_id>/pdf")
@login_required
def bon_pdf(of_id):
    of = get_of_or_403(of_id)
    stations = of.stations.all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle("titre", parent=styles["Heading1"],
                                 fontSize=16, textColor=colors.HexColor("#0C447C"),
                                 alignment=TA_CENTER, spaceAfter=6)
    sous_titre_style = ParagraphStyle("sous_titre", parent=styles["Normal"],
                                      fontSize=10, textColor=colors.HexColor("#5F5E5A"),
                                      alignment=TA_CENTER, spaceAfter=14)
    label_style = ParagraphStyle("label", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#888780"))
    valeur_style = ParagraphStyle("valeur", parent=styles["Normal"],
                                  fontSize=10, textColor=colors.HexColor("#2C2C2A"))

    elements = []

    # En-tête
    elements.append(Paragraph("BON DE PRÉPARATION D'ENCRE", titre_style))
    elements.append(Paragraph(f"Triax Ink Management — Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", sous_titre_style))

    # Infos OF
    type_tirage_label = "Bobine" if of.type_tirage == "BOBINE" else "Feuille à feuille"
    if of.type_tirage == "BOBINE":
        format_str = f"Laize : {of.laize_m} m — Métrage : {of.metrage_m} m"
    else:
        format_str = f"{of.hauteur_m*100:.0f} cm × {of.largeur_m*100:.0f} cm — {of.nb_tirages} tirages"

    info_data = [
        ["N° OF", of.reference_of, "Produit", of.nom_produit or "—"],
        ["Client", of.client.nom, "Date", of.date_of.strftime("%d/%m/%Y")],
        ["Type", type_tirage_label, "Format / Tirage", format_str],
        ["Surface totale", f"{of.surface_m2} m²", "Marge appliquée", f"{of.marge_appliquee_pct}%"],
    ]

    info_table = Table(info_data, colWidths=[3*cm, 6*cm, 3.5*cm, 4.5*cm])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#5F5E5A")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#5F5E5A")),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#F1EFE8"), colors.white]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D1C7")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D3D1C7")),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*cm))

    # Tableau des stations
    headers = ["Couleur", "Encre", "Anilox", "Vol.\ncm³/m²", "Coeff.", "Couv.\n%",
               "Masse nette\n(kg)", "Masse recommandée\n(kg)"]
    rows = [headers]

    total_nette = 0
    total_marge = 0

    for s in stations:
        rows.append([
            s.nom_couleur,
            s.encre_label,
            s.anilox_label,
            f"{s.anilox_volume:.2f}",
            f"{s.anilox_coeff:.2f}",
            f"{s.taux_couverture_pct:.1f}%" if s.taux_couverture_pct is not None else "—",
            f"{s.masse_nette_kg:.3f}" if s.masse_nette_kg is not None else "—",
            f"{s.masse_avec_marge_kg:.3f}" if s.masse_avec_marge_kg is not None else "—",
        ])
        total_nette += s.masse_nette_kg or 0
        total_marge += s.masse_avec_marge_kg or 0

    rows.append(["TOTAL", "", "", "", "", "",
                 f"{total_nette:.3f} kg", f"{total_marge:.3f} kg"])

    col_widths = [2.5*cm, 4*cm, 2.5*cm, 1.8*cm, 1.8*cm, 1.5*cm, 2.5*cm, 3*cm]
    tableau = Table(rows, colWidths=col_widths, repeatRows=1)
    tableau.setStyle(TableStyle([
        # En-tête
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0C447C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        # Corps
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F1EFE8")]),
        ("ALIGN", (3, 1), (-1, -1), "CENTER"),
        # Ligne total
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E6F1FB")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#0C447C")),
        # Colonne masse recommandée
        ("FONTNAME", (-1, 1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (-1, 1), (-1, -2), colors.HexColor("#185FA5")),
        # Bordures
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D1C7")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D3D1C7")),
        ("PADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(tableau)
    elements.append(Spacer(1, 0.4*cm))

    # Note de bas
    elements.append(Paragraph(
        f"Source des taux de couverture : {'Analyse IA (PDF)' if of.taux_source == 'IA' else 'Saisie manuelle'} — "
        f"Marge de sécurité : {of.marge_appliquee_pct}%",
        ParagraphStyle("note", parent=styles["Normal"], fontSize=8,
                       textColor=colors.HexColor("#888780"), alignment=TA_LEFT)
    ))

    doc.build(elements)
    buffer.seek(0)
    filename = f"bon_encre_{of.reference_of}_{of.date_of.strftime('%Y%m%d')}.pdf"
    return send_file(buffer, mimetype="application/pdf",
                     as_attachment=True, download_name=filename)


@export_bp.route("/of/<int:of_id>/excel")
@login_required
def bon_excel(of_id):
    of = get_of_or_403(of_id)
    stations = of.stations.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"OF {of.reference_of}"

    # Styles
    bleu_fonce = "0C447C"
    bleu_clair = "E6F1FB"
    gris_clair = "F1EFE8"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=bleu_fonce)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    total_font = Font(bold=True, color=bleu_fonce, size=10)
    total_fill = PatternFill("solid", fgColor=bleu_clair)

    thin = Side(style="thin", color="D3D1C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Ligne titre
    ws.merge_cells("A1:H1")
    ws["A1"] = f"BON DE PRÉPARATION — OF {of.reference_of}"
    ws["A1"].font = Font(bold=True, size=14, color=bleu_fonce)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Infos OF
    infos = [
        ("Client", of.client.nom), ("Produit", of.nom_produit or "—"),
        ("Date OF", of.date_of.strftime("%d/%m/%Y")),
        ("Surface", f"{of.surface_m2} m²"),
        ("Marge sécurité", f"{of.marge_appliquee_pct}%"),
        ("Source couverture", "Analyse IA (PDF)" if of.taux_source == "IA" else "Saisie manuelle"),
    ]
    for row_i, (label, val) in enumerate(infos, start=2):
        ws.cell(row=row_i, column=1, value=label).font = Font(bold=True, color="5F5E5A", size=9)
        ws.cell(row=row_i, column=2, value=val).font = Font(size=9)

    # En-tête tableau
    header_row = len(infos) + 3
    headers = ["Couleur", "Encre", "Anilox", "Vol. (cm³/m²)", "Coeff. transfert",
               "Couverture (%)", "Masse nette (kg)", "Masse recommandée (kg)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[header_row].height = 30

    total_nette = total_marge = 0
    for i, s in enumerate(stations):
        r = header_row + 1 + i
        fill = PatternFill("solid", fgColor=gris_clair) if i % 2 == 0 else None
        row_data = [
            s.nom_couleur, s.encre_label, s.anilox_label,
            s.anilox_volume, s.anilox_coeff,
            s.taux_couverture_pct,
            s.masse_nette_kg, s.masse_avec_marge_kg,
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            cell.font = Font(size=9, bold=(col == 8), color=(bleu_fonce if col == 8 else "000000"))
            cell.alignment = Alignment(horizontal="center" if col > 3 else "left")
            if fill:
                cell.fill = fill
        total_nette += s.masse_nette_kg or 0
        total_marge += s.masse_avec_marge_kg or 0

    # Ligne total
    tr = header_row + 1 + len(stations)
    ws.cell(row=tr, column=1, value="TOTAL").font = total_font
    ws.cell(row=tr, column=7, value=round(total_nette, 3)).font = total_font
    ws.cell(row=tr, column=8, value=round(total_marge, 3)).font = total_font
    for col in range(1, 9):
        ws.cell(row=tr, column=col).fill = total_fill
        ws.cell(row=tr, column=col).border = border

    # Largeurs colonnes
    for col, width in zip("ABCDEFGH", [14, 22, 14, 14, 16, 15, 18, 22]):
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"bon_encre_{of.reference_of}_{of.date_of.strftime('%Y%m%d')}.xlsx"
    return send_file(buffer,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)
